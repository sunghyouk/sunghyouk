import os
import pathlib

import openpyxl


def read_files(dir_path, filename):

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    xl_filename = filename + '.xlsx'
    ws.cell(1, 1).value = 'Number'
    ws.cell(1, 2).value = 'Name of Book'
    wb.save('./' + xl_filename)

    path = pathlib.Path(dir_path)
    print(path)

    files = os.listdir(path)
    files_list_pdf = [file for file in files if file.endswith(".pdf")]
    print(files_list_pdf)

    row_cnt = 2
    for file in files_list_pdf:
        if file.is_file():
            print(file.name)
            ws.cell(row_cnt, 1).value = row_cnt - 1
            ws.cell(row_cnt, 2).value = file.name
            row_cnt += 1
    wb.save('./' + xl_filename)
    wb.close()


directory = r'/Users/sunghyouk/Library/CloudStorage/OneDrive-개인/CS_library'
rfilename = 'CS_book_list'

read_files(directory, rfilename)
